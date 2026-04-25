import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule
from datetime import datetime
import re
import io
import traceback

def _safe_value(v):
    """将单元格值转换为 openpyxl 可写入的类型，NaN 转为 ''"""
    if isinstance(v, float) and pd.isna(v):
        return ''
    if v is None or (isinstance(v, float) and np.isnan(v)):
        return ''
    return v

def read_excel_file(file_obj):
    df = pd.read_excel(file_obj, header=None)
    first_val = str(df.iloc[1, 0]) if pd.notna(df.iloc[1, 0]) else ''
    
    if first_val.isdigit() and len(first_val) > 5:
        data = df.iloc[1:].copy()
        data.columns = [str(c).strip() for c in df.iloc[0]]
    else:
        headers = df.iloc[0].tolist()
        score_cn = df.iloc[1].tolist()
        data = df.iloc[2:].copy()
        cols = []
        for i, h in enumerate(headers):
            if i < len(score_cn) and pd.notna(score_cn[i]):
                cname = str(score_cn[i])
                cname = re.sub(r'（[^（）]*分[^（）]*）', '', cname)
                cname = re.sub(r'\([^()]*分[^()]*\)', '', cname)
                cname = cname.strip()
                cols.append(cname)
            elif pd.notna(h):
                cols.append(str(h))
            else:
                cols.append(f"Unnamed_{i}")
        data.columns = cols

    data.columns = [str(c).strip() for c in data.columns]
    data = data.loc[:, ~data.columns.duplicated()]
    data = data.dropna(subset=['学号']).reset_index(drop=True)

    if '班级' in data.columns:
        data['班级'] = data['班级'].astype(str).str.strip()
        data['班级'] = data['班级'].fillna('未知班级')
    return data

def classify_by_subject(data):
    answer_idxs = set()
    cols = list(data.columns)
    for i, c in enumerate(cols):
        if '选' in str(c) and i+1 < len(cols):
            nxt = cols[i+1]
            if pd.isna(nxt) or str(nxt).strip() == '' or 'Unnamed' in str(nxt):
                answer_idxs.add(i+1)

    score_cols = []
    for i, c in enumerate(cols):
        if c in ['学号','考号','姓名','班级','学校','全卷','语文','1卷','2卷']:
            continue
        if i in answer_idxs:
            continue
        try:
            ser = pd.to_numeric(data[c], errors='coerce')
            if ser.notna().any():
                data[c] = ser
                score_cols.append(c)
        except:
            pass

    def natural_key(s):
        if '25' in s: return (999,0,0,0)
        nums = re.findall(r'\d+', s)
        if nums:
            lst = [int(x) for x in nums]
            while len(lst)<3: lst.append(0)
            return (0, lst[0], lst[1], lst[2])
        return (2,0,0,0)
    score_cols = sorted(score_cols, key=natural_key)
    cn_cols = [c for c in score_cols if '语' in c]
    if not cn_cols:
        cn_cols = score_cols.copy()
    return cn_cols

def calculate_class_averages(data, score_cols=None):
    if score_cols is None:
        score_cols = classify_by_subject(data)
    else:
        score_cols = [c for c in score_cols if c in data.columns]

    cn = [c for c in score_cols if '语' in c]
    if not cn:
        cn = score_cols

    def natural_key(s):
        if '25' in s: return (999,0,0,0)
        nums = re.findall(r'\d+', s)
        if nums:
            lst = [int(x) for x in nums]
            while len(lst)<3: lst.append(0)
            return (0, lst[0], lst[1], lst[2])
        return (2,0,0,0)
    score_cols = sorted(score_cols, key=natural_key)

    class_avg = data.groupby('班级')[score_cols].mean().round(2)
    class_avg = class_avg.sort_index()
    total_avg = data[score_cols].mean().round(2)
    total_avg = pd.DataFrame([total_avg], index=['总平均值'])
    all_avg = pd.concat([total_avg, class_avg])
    all_avg.insert(0, '全卷', all_avg[score_cols].sum(axis=1).round(2))
    if cn:
        all_avg.insert(1, '语文', all_avg[cn].sum(axis=1).round(2))
    return all_avg

def generate_cn_analysis(data, cn_cols):
    seen = set()
    cn_cols_u = []
    for c in cn_cols:
        if c in data.columns and c not in seen:
            seen.add(c)
            cn_cols_u.append(c)
    class_an = data.groupby('班级')[cn_cols_u].mean().round(2)
    class_an = class_an.sort_index()
    total_an = data[cn_cols_u].mean().round(2)
    total_an = pd.DataFrame([total_an], index=['总平均值'])
    all_an = pd.concat([total_an, class_an])
    all_an.insert(0, '语文成绩', all_an[cn_cols_u].sum(axis=1).round(2))
    return all_an

def generate_subject_analysis(data, cn_cols, subject_name='语文'):
    return generate_cn_analysis(data, cn_cols)

def add_color_scale(ws, start_col, end_col, start_row, end_row):
    try:
        rng = f'{start_col}{start_row}:{end_col}{end_row}'
        rule = ColorScaleRule(
            start_type='min', start_color='FFF16B61',
            mid_type='percentile', mid_value=50, mid_color='FFF7E98F',
            end_type='max', end_color='FF64BC7B'
        )
        ws.conditional_formatting.add(rng, rule)
    except Exception as e:
        print(f"色阶添加失败（可忽略）: {e}")

def _create_workbook(all_avg, cn_analysis, data):
    wb = Workbook()
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # ---------- 工作表1：班级平均分 ----------
    ws1 = wb.active
    ws1.title = "班级平均分"
    ws1.append(['班级'] + list(all_avg.columns))
    for idx, row in all_avg.iterrows():
        ws1.append([idx] + [_safe_value(row[c]) for c in all_avg.columns])

    for cell in ws1[1]:
        cell.fill = PatternFill(start_color="FFCCCCCC", end_color="FFCCCCCC", fill_type="solid")
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in ws1.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
            if cell.column == 1:
                cell.font = Font(bold=True)

    # ---------- 工作表2：语文学科分析 ----------
    ws2 = wb.create_sheet(title="语文学科分析")
    ws2.append(['班级'] + list(cn_analysis.columns))
    for idx, row in cn_analysis.iterrows():
        ws2.append([idx] + [_safe_value(row[c]) for c in cn_analysis.columns])

    for cell in ws2[1]:
        cell.fill = PatternFill(start_color="FFCCCCCC", end_color="FFCCCCCC", fill_type="solid")
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in ws2.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
            if cell.column == 1:
                cell.font = Font(bold=True)

    if cn_analysis.shape[0] > 0 and cn_analysis.shape[1] > 0:
        start_row = 2
        end_row = start_row + len(cn_analysis) - 1
        for col_idx in range(cn_analysis.shape[1]):
            col_num = col_idx + 2
            col_letter = ''
            while col_num > 0:
                col_num, rem = divmod(col_num - 1, 26)
                col_letter = chr(65 + rem) + col_letter
            add_color_scale(ws2, col_letter, col_letter, start_row, end_row)

    for ws in [ws1, ws2]:
        for column in ws.columns:
            ws.column_dimensions[column[0].column_letter].width = 8
        for row in ws.iter_rows():
            ws.row_dimensions[row[0].row].height = 14

    # ---------- 各班小题工作表 ----------
    # 收集小题列（过滤空列头、'nan' 字符串、Unnamed）
    score_cols = []
    for col in data.columns:
        if col in ['学号','考号','姓名','班级','学校','全卷','语文','1卷','2卷']:
            continue
        if pd.isna(col) or str(col).strip() == '' or 'Unnamed' in str(col):
            continue
        if str(col).strip().lower() == 'nan':
            continue
        score_cols.append(col)

    # 去重
    seen = set()
    score_cols_u = []
    for c in score_cols:
        if c not in seen:
            seen.add(c)
            score_cols_u.append(c)
    score_cols = score_cols_u

    # 格式化列名：整数浮点数如 1.0 显示为 1
    def _format_col(col):
        s = str(col).strip()
        try:
            f = float(s)
            if f == int(f):
                return str(int(f))
            else:
                return s
        except ValueError:
            return s

    score_cols_display = [_format_col(c) for c in score_cols]

    # 自然排序
    def natural_key(s):
        if '25' in s:
            return (999, 0, 0, 0)
        nums = re.findall(r'\d+', s)
        if nums:
            lst = [int(x) for x in nums]
            while len(lst) < 3:
                lst.append(0)
            return (0, lst[0], lst[1], lst[2])
        return (2, 0, 0, 0)

    # 对原始列名排序，显示列名跟随
    original_sorted = sorted(score_cols, key=natural_key)
    display_sorted = []
    for orig in original_sorted:
        idx = score_cols.index(orig)
        display_sorted.append(score_cols_display[idx])

    base_cols = ['姓名', '班级', '全卷', '语文']
    classes = sorted(data['班级'].astype(str).str.strip().unique())

    for cls in classes:
        c_data = data[data['班级'] == cls].copy()
        if c_data.empty:
            continue
        c_data = c_data.loc[:, ~c_data.columns.duplicated()]

        header = [str(c) for c in base_cols + display_sorted]
        try:
            sheet_name = str(cls)[:31]
            for ch in ['[', ']', ':', '*', '?', '/', '\\']:
                sheet_name = sheet_name.replace(ch, '_')
            ws_class = wb.create_sheet(title=sheet_name)
        except Exception as e:
            print(f"创建 sheet '{cls}' 失败: {e}")
            continue

        ws_class.append(header)

        for _, row in c_data.iterrows():
            row_dict = row.to_dict()
            data_row = []
            for col in base_cols:
                if col == '班级':
                    data_row.append(cls)
                else:
                    data_row.append(_safe_value(row_dict.get(col, '')))
            for col in original_sorted:
                data_row.append(_safe_value(row_dict.get(col, '')))
            ws_class.append(data_row)

        # 样式
        for cell in ws_class[1]:
            cell.fill = PatternFill(start_color="FFCCCCCC", end_color="FFCCCCCC", fill_type="solid")
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for row in ws_class.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border

        # 为每个小题列添加色阶（替换零分标红）
        base_len = len(base_cols)
        start_row_cls = 2
        end_row_cls = ws_class.max_row
        if end_row_cls >= start_row_cls and len(original_sorted) > 0:
            for c_idx in range(base_len, base_len + len(original_sorted)):
                col_num = c_idx + 1          # Excel 列号从1开始
                col_letter = ''
                while col_num > 0:
                    col_num, rem = divmod(col_num - 1, 26)
                    col_letter = chr(65 + rem) + col_letter
                add_color_scale(ws_class, col_letter, col_letter, start_row_cls, end_row_cls)

        # 列宽行高
        for column in ws_class.columns:
            ws_class.column_dimensions[column[0].column_letter].width = 8
        for row in ws_class.iter_rows():
            ws_class.row_dimensions[row[0].row].height = 14

    return wb

def save_results_to_excel(all_avg, cn_analysis, output_file, data):
    wb = _create_workbook(all_avg, cn_analysis, data)
    wb.save(output_file)

def save_results_to_excel_bytes(all_avg, cn_analysis, data):
    wb = _create_workbook(all_avg, cn_analysis, data)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def main():
    print("正在读取数据...")
    data = read_excel_file('data.xls')
    print(f"成功读取 {len(data)} 条记录")
    print("正在计算班级平均分...")
    class_averages = calculate_class_averages(data)
    print("班级平均分计算完成")
    print("正在分类数据...")
    cn_cols = classify_by_subject(data)
    print(f"语文学科包含 {len(cn_cols)} 个小题")
    print("正在生成语文学科分析...")
    cn_analysis = generate_cn_analysis(data, cn_cols)
    print("正在保存结果...")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = f'分析结果_{timestamp}.xlsx'
    save_results_to_excel(class_averages, cn_analysis, output_file, data)
    print(f"分析结果已保存到 '{output_file}'")

if __name__ == "__main__":
    main()
