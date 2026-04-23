import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime

# 读取Excel文件 - 支持文件路径或文件对象
def read_excel_file(file_obj):
    # 读取原始数据
    df = pd.read_excel(file_obj)
    
    # 处理表头
    # 第一行是字段名，第二行是分数说明
    headers = df.iloc[0].tolist()
    score_chinese = df.iloc[1].tolist()
    
    # 提取实际数据（从第三行开始）
    data = df.iloc[2:].copy()
    
    # 设置列名
    columns = []
    for i, col in enumerate(headers):
        # 优先使用第二行的值（分数说明），因为第一行可能有重复的"1卷"等
        if i < len(score_chinese) and pd.notna(score_chinese[i]):
            col_name = str(score_chinese[i])
            # 去掉分数部分，只保留名称
            # 使用正则表达式删除所有包含"分"字的括号及其内容
            import re
            # 匹配全角括号中的分数部分
            col_name = re.sub(r'（[^（）]*分[^（）]*）', '', col_name)
            # 匹配半角括号中的分数部分
            col_name = re.sub(r'\([^()]*分[^()]*\)', '', col_name)
            # 去除可能的多余空格
            col_name = col_name.strip()
            columns.append(col_name)
        elif pd.notna(col):
            columns.append(str(col))
        else:
            columns.append(f"Unnamed_{i}")
    
    data.columns = columns
    
    # 清理数据
    data = data.dropna(subset=['学号']).reset_index(drop=True)
    
    # 将"技术"列重命名为"全卷"
    if '语文' in data.columns:
        data = data.rename(columns={'语文': '全卷'})
    
    return data

# 计算班级平均分
def calculate_class_averages(data):
    # 提取需要计算平均分的列（只包含分数列，不包含答案列和试卷分类列）
    score_columns = []
    chinese_cols = []
    
    for col in data.columns:
        if col not in ['学号', '考号', '姓名', '行政班级', '学校', '全卷', '语文', '1卷', '2卷']:
            # 只选择不包含"答案"的列
            if '答案' not in col:
                try:
                    data[col] = pd.to_numeric(data[col], errors='coerce')
                    score_columns.append(col)
                    
                    chinese_cols.append(col)
                    """
                    if '选' in col:
                        info_cols.append(col)
                    elif '通' in col:
                        general_cols.append(col)
                    """
                except:
                    pass
    
    # 按照题目编号的自然顺序排列列
    import re
    
    def natural_sort_key(s):
        # 提取题目编号
        if '25' in s:
            return (999, 0, 0, 0)  # 999表示你不会写的作文，放在最后
        else:
            # 提取所有数字
            numbers = re.findall(r'\d+', s)
            if numbers:
                # 将数字转换为整数，确保数值排序而非字符串排序
                num_list = [int(n) for n in numbers]
                # 确保至少有3个数字，用于比较
                while len(num_list) < 3:
                    num_list.append(0)
                return (0, num_list[0], num_list[1], num_list[2])  # 0表示正常的语文题
        return (2, 0, 0, 0)  # 其他情况
    
    score_columns = sorted(score_columns, key=natural_sort_key)
    
    # 按班级分组计算平均分
    class_avg = data.groupby('行政班级')[score_columns].mean().round(2)
    
    # 按班级名称升序排序
    class_avg = class_avg.sort_index()
    
    # 计算全年级平均分
    total_avg = data[score_columns].mean().round(2)
    total_avg = pd.DataFrame([total_avg], index=['总平均值'])
    
    # 合并结果
    all_avg = pd.concat([total_avg, class_avg])
    
    # 添加汇总列（放在最前面）
    # 全卷（所有分数的总和）
    all_avg.insert(0, '全卷', all_avg[score_columns].sum(axis=1).round(2))
    # 语文（语文学科分数的总和）
    if chinese_cols:
        all_avg.insert(1, '语文', all_avg[chinese_cols].sum(axis=1).round(2))
    
    return all_avg

# 按学科分类数据
def classify_by_subject(data):
    # 提取需要分析的列（只包含分数列，不包含答案列和试卷分类列）
    score_columns = []
    excluded_cols = []
    non_numeric_cols = []
    
    print("\n调试信息：所有列的处理情况")
    print(f"总列数: {len(data.columns)}")
    print("\n列名列表:")
    for i, col in enumerate(data.columns):
        print(f"{i+1}. {col}")
    
    print("\n详细处理过程:")
    for col in data.columns:
        if col in ['学号', '考号', '姓名', '行政班级', '学校', '全卷', '信息', '通用', '1卷', '2卷']:
            excluded_cols.append(col)
            print(f"  排除列: {col} (非分数列)")
        elif '答案' in col:
            excluded_cols.append(col)
            print(f"  排除列: {col} (答案列)")
        else:
            try:
                # 尝试转换为数值类型
                converted = pd.to_numeric(data[col], errors='coerce')
                # 检查是否有非NaN值
                if not converted.isnull().all():
                    data[col] = converted
                    score_columns.append(col)
                    print(f"  保留列: {col} (数值列)")
                else:
                    non_numeric_cols.append(col)
                    print(f"  排除列: {col} (全部为非数值)")
            except Exception as e:
                non_numeric_cols.append(col)
                print(f"  排除列: {col} (转换失败: {str(e)})")
    
    print(f"\n处理结果:")
    print(f"  保留的分数列: {len(score_columns)}")
    print(f"  排除的列: {len(excluded_cols)}")
    print(f"  非数值列: {len(non_numeric_cols)}")
    
    # 按照题目编号的自然顺序排列列
    import re
    def natural_sort_key(s):
        # 提取题目编号
        if '25' in s:
            return (999, 0, 0, 0)  # 25表示作文，放在最后
        # 检查是否包含"信"或"通"字
        else:
            # 提取所有数字
            numbers = re.findall(r'\d+', s)
            if numbers:
                # 将数字转换为整数，确保数值排序而非字符串排序
                num_list = [int(n) for n in numbers]
                # 确保至少有3个数字，用于比较
                while len(num_list) < 3:
                    num_list.append(0)
                return (0, num_list[0], num_list[1], num_list[2])  # 0表示信息题
    
    score_columns = sorted(score_columns, key=natural_sort_key)
    
    # 分类
    chinese_cols = [col for col in score_columns]
    
    return chinese_cols

# 生成信息学科分析
def generate_chinese_analysis(data, chinese_cols):
    # 按班级分组计算
    class_analysis = data.groupby('行政班级')[chinese_cols].mean().round(2)
    
    # 按班级名称升序排序
    class_analysis = class_analysis.sort_index()
    
    # 计算全年级数据
    total_analysis = data[chinese_cols].mean().round(2)
    total_analysis = pd.DataFrame([total_analysis], index=['总平均值'])
    
    # 合并结果
    all_analysis = pd.concat([total_analysis, class_analysis])
    
    # 添加信息成绩列（放在最前面）
    all_analysis.insert(0, '语文成绩', all_analysis[chinese_cols].sum(axis=1).round(2))
    
    return all_analysis


# 添加色阶效果
from openpyxl.formatting.rule import ColorScaleRule

def add_color_scale(ws, start_col, end_col, start_row, end_row):
    # 为指定范围的单元格添加色阶效果
    # 红色表示较低值，绿色表示较高值
    # 使用用户提供的具体颜色值
    try:
        # 生成正确格式的范围字符串
        range_str = f'{start_col}{start_row}:{end_col}{end_row}'
        
        # 使用用户提供的颜色值
        # 红色表示较低值，黄色表示中间值，绿色表示较高值
        color_scale_rule = ColorScaleRule(
            start_type='min', start_color='FFF16B61',  # 深橙红
            mid_type='percentile', mid_value=50, mid_color='FFF7E98F',  # 淡黄色
            end_type='max', end_color='FF64BC7B'  # 深草绿
        )
        ws.conditional_formatting.add(range_str, color_scale_rule)
    except Exception as e:
        print(f"添加色阶效果时出错: {str(e)}")
        print(f"范围: {start_col}{start_row}:{end_col}{end_row}")

import io

# 保存结果到Excel文件
def save_results_to_excel(all_avg, chinese_analysis,  output_file, data):
    wb = _create_workbook(all_avg, chinese_analysis,  data)
    wb.save(output_file)

# 保存结果到字节流（用于Web应用）
def save_results_to_excel_bytes(all_avg, chinese_analysis,  data):
    wb = _create_workbook(all_avg, chinese_analysis,  data)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# 创建工作簿的内部函数
def _create_workbook(all_avg, chinese_analysis,  data):
    wb = Workbook()
        
    # ws2工作表：信息学科分析
    ws2 = wb.create_sheet(title="语文学科分析")
    
    # 写入数据
    # 先写入表头
    header_row = ['班级'] + list(chinese_analysis.columns)
    ws2.append(header_row)
    
    # 然后写入数据
    for idx, row in chinese_analysis.iterrows():
        data_row = [idx] + list(row.values)
        ws2.append(data_row)
    
    # 设置样式
    for row in ws2.iter_rows(min_row=1, max_row=1):
        for cell in row:
            cell.fill = PatternFill(start_color="FFCCCCCC", end_color="FFCCCCCC", fill_type="solid")
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # 居中对齐所有数据
    for row in ws2.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # 为信息学科分析添加色阶效果（按列比较）
    if chinese_analysis_analysis.shape[0] > 0 and chinese_analysis.shape[1] > 0:
        start_row = 2
        end_row = start_row + len(chinese_analysis) - 1
        # 为每一列单独添加色阶效果
        for col_idx in range(chinese_analysis.shape[1]):
            # 正确计算Excel列字母（从B列开始）
            col_num = col_idx + 2  # B列是第2列
            col_letter = ''
            while col_num > 0:
                col_num, remainder = divmod(col_num - 1, 26)
                col_letter = chr(65 + remainder) + col_letter
            add_color_scale(ws2, col_letter, col_letter, start_row, end_row)
    
    
    # 调整列宽和行高，设置样式
    from openpyxl.styles import Border, Side
    
    # 定义边框样式
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for ws in [ws2]:
        # 设置列宽为8字符
        for column in ws.columns:
            column_letter = column[0].column_letter
            ws.column_dimensions[column_letter].width = 8
        
        # 设置行高为14磅
        for row in ws.iter_rows():
            ws.row_dimensions[row[0].row].height = 14
        
        # 设置表头加粗
        for cell in ws[1]:
            cell.font = Font(bold=True)
        
        # 设置首列（班级列）加粗
        for row in ws.iter_rows(min_row=2):
            row[0].font = Font(bold=True)
        
        # 设置所有单元格边框
        for row in ws.iter_rows():
            for cell in row:
                cell.border = border
    
    # 为每个班级创建单独的sheet表
    # 获取所有班级列表并按升序排序
    classes = data['行政班级'].unique()
    classes = sorted(classes)
    
    for cls in classes:
        # 筛选该班级的数据
        class_data = data[data['行政班级'] == cls].copy()
        
        # 创建班级工作表
        ws_class = wb.create_sheet(title=f"{cls}")
        
        # 参考calculate_class_averages函数的实现，动态提取通用小题列
        # 基础列（在最前方增加姓名列）
        base_columns = ['姓名', '班级', '全卷', '语文']
        
        # 提取通用小题列
        chinese_columns = []
        excluded_cols = ['学号', '考号', '姓名', '行政班级', '学校', '全卷', '语文', '1卷', '2卷']
        
        for col in class_data.columns:
            if col not in excluded_cols:
                # 只选择不包含"答案"的列
                if '答案' not in col:
                    try:
                        # 尝试转换为数值类型
                        pd.to_numeric(class_data[col], errors='coerce')
                        # 只选择包含"通"的列
                        chinese_columns.append(col)
                    except:
                        pass
        
        # 按照题目编号的自然顺序排列列（参考calculate_class_averages函数）
        import re
        def natural_sort_key(s):
            # 提取题目编号
            # 首先检查是否包含"草图"字，放在最后
            if '25' in s:
                return (999, 0, 0, 0)  # 999表示草图，放在最后
            # 检查是否包含"通"字
            else:
                # 提取所有数字
                numbers = re.findall(r'\d+', s)
                if numbers:
                    # 将数字转换为整数，确保数值排序而非字符串排序
                    num_list = [int(n) for n in numbers]
                    # 确保至少有3个数字，用于比较
                    while len(num_list) < 3:
                        num_list.append(0)
                    return (1, num_list[0], num_list[1], num_list[2])  # 1表示通
            return (2, 0, 0, 0)  # 其他情况
        
        chinese_columns = sorted(chinese_columns, key=natural_sort_key)
        
        # 构建完整表头
        header_row = base_columns + chinese_columns
        
        # 确保班级列存在
        if '行政班级' in class_data.columns:
            class_data['班级'] = class_data['行政班级']
        
        # 确保姓名列存在
        if '姓名' not in class_data.columns:
            class_data['姓名'] = ''
        
        # 确保基础列存在并获取正确值
        # 全卷列
        if '全卷' not in class_data.columns:
            # 尝试从其他可能的列名获取
            if '语文' in class_data.columns:
                class_data['全卷'] = class_data['语文']
            else:
                class_data['全卷'] = 0
        
        """
        # 信息列
        if '信息' not in class_data.columns:
            # 尝试计算每个学生的信息学科分数
            chinese_cols = [col for col in class_data.columns if '信' in col and '答案' not in col]
            if chinese_cols:
                # 计算每个学生的信息学科总分
                class_data['信息'] = 0
                for col in chinese_cols:
                    try:
                        class_data['信息'] += class_data[col].astype(float)
                    except:
                        pass
            else:
                class_data['信息'] = 0
        """
        
        # 写入表头
        ws_class.append(header_row)
        
        # 写入数据
        for _, row in class_data.iterrows():
            data_row = []
            # 添加基础列数据
            for col in base_columns:
                if col in row:
                    data_row.append(row[col])
                else:
                    data_row.append('')
            # 添加通用小题数据
            for col in general_columns:
                if col in row:
                    data_row.append(row[col])
                else:
                    data_row.append('')
            ws_class.append(data_row)
        
        # 设置样式
        # 表头样式
        for row in ws_class.iter_rows(min_row=1, max_row=1):
            for cell in row:
                cell.fill = PatternFill(start_color="FFCCCCCC", end_color="FFCCCCCC", fill_type="solid")
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # 数据样式
        for row in ws_class.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border
        
        # 标记0分的通用小题单元格
        # 基础列数量（姓名、班级、全卷、信息、通用）
        base_col_count = len(base_columns)
        # 遍历数据行
        for row_idx in range(2, ws_class.max_row + 1):
            # 遍历通用小题列
            for col_idx in range(base_col_count, base_col_count + len(chinese_columns)):
                cell = ws_class.cell(row=row_idx, column=col_idx + 1)  # Excel列从1开始
                # 检查单元格值是否为0
                if cell.value == 0:
                    # 设置背景颜色为#ffc7ce
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        # 设置列宽和行高
        for column in ws_class.columns:
            column_letter = column[0].column_letter
            ws_class.column_dimensions[column_letter].width = 8
        
        for row in ws_class.iter_rows():
            ws_class.row_dimensions[row[0].row].height = 14
    
    return wb

# 主函数
def main():
    # 读取数据
    print("正在读取数据...")
    data = read_excel_file('data.xls')
    print(f"成功读取 {len(data)} 条记录")
    
    # 计算班级平均分
    print("正在计算班级平均分...")
    class_averages = calculate_class_averages(data)
    print("班级平均分计算完成")
    
    # 分类数据
    print("正在分类数据...")
    chinese_cols = classify_by_subject(data)
    print(f"语文学科包含 {len(chinese_cols)} 个小题")
    print(f"识别到的语文类题目表头: {chinese_cols}")
    
    # 生成信息学科分析
    print("正在生成语文学科分析...")
    chinese_analysis = generate_chinese_analysis(data, chinese_cols)
    
    # 保存结果
    print("正在保存结果...")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = f'分析结果_{timestamp}.xlsx'
    save_results_to_excel(class_averages, chinese_analysis,  output_file, data)
    print(f"分析结果已保存到 '{output_file}'")

if __name__ == "__main__":
    main()
