import pandas as pd
import numpy as np
import re
import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime


# 读取Excel文件 - 支持文件路径或文件对象
def read_excel_file(file_obj):
    # 读取原始数据
    df = pd.read_excel(file_obj)
    
    # 处理表头
    # 第一行是字段名，第二行是分数说明
    headers = df.iloc[0].tolist()
    score_chinese = df.iloc[1].tolist()
    
    # 提取实际数据（从第三行开始）
    data = df.iloc[2:].copy()
    
    # 设置列名
    columns = []
    for i, col in enumerate(headers):
        # 优先使用第二行的值（分数说明），因为第一行可能有重复的"1卷"等
        if i < len(score_chinese) and pd.notna(score_chinese[i]):
            col_name = str(score_chinese[i])
            # 去掉分数部分，只保留名称
            # 匹配全角括号中的分数部分
            col_name = re.sub(r'（[^（）]*分[^（）]*）', '', col_name)
            # 匹配半角括号中的分数部分
            col_name = re.sub(r'\([^()]*分[^()]*\)', '', col_name)
            # 去除可能的多余空格
            col_name = col_name.strip()
            columns.append(col_name)
        elif pd.notna(col):
            columns.append(str(col))
        else:
            columns.append(f"Unnamed_{i}")
    
    data.columns = columns
    
    # 清理数据
    data = data.dropna(subset=['学号']).reset_index(drop=True)
    
    # 将"语文"列重命名为"全卷"
    if '语文' in data.columns:
        data = data.rename(columns={'语文': '全卷'})
    
    return data


# 自然排序函数
def natural_sort_key(s):
    """按题目编号自然排序，作文题（25题）放在最后"""
    # 精确匹配：如果列名中包含独立的25（作为题号），例如 "25题" 或 "25、..."，放在最后
    # 简单用正则查找 \b25\b 更准确，但考虑到中文数字混合，我们使用：若以 25 开头或包含非数字分隔的 25
    if re.search(r'(?<!\d)25(?!\d)', s):
        return (999, 0, 0, 0)
    numbers = re.findall(r'\d+', s)
    if numbers:
        num_list = [int(n) for n in numbers]
        while len(num_list) < 3:
            num_list.append(0)
        return (0, num_list[0], num_list[1], num_list[2])
    return (2, 0, 0, 0)


# 提取数值型小题列（排除答案列和固定信息列）
def get_score_columns(data, subject_prefix=None):
    """
    从数据中提取可以作为分数的列。
    排除：学号、考号、姓名、行政班级、学校、全卷、语文、1卷、2卷 及所有包含“答案”的列。
    保留可转换为数值且至少有一个非空值的列。
    """
    exclude_base = {'学号', '考号', '姓名', '行政班级', '学校', '全卷', '语文', '1卷', '2卷'}
    score_cols = []
    
    for col in data.columns:
        if col in exclude_base:
            continue
        if '答案' in col:
            continue
        # 检查是否可视为数值列
        numeric_col = pd.to_numeric(data[col], errors='coerce')
        if numeric_col.notna().any():  # 至少有一个有效数值
            data[col] = numeric_col   # 原地转为数值（便于后续计算）
            score_cols.append(col)
    
    # 按题目编号自然顺序排序
    score_cols = sorted(score_cols, key=natural_sort_key)
    return score_cols


# 计算班级平均分
def calculate_class_averages(data, score_columns):
    # 按班级分组计算平均分
    class_avg = data.groupby('行政班级')[score_columns].mean().round(2)
    class_avg = class_avg.sort_index()
    
    # 计算全年级平均分
    total_avg = data[score_columns].mean().round(2)
    total_avg = pd.DataFrame([total_avg], index=['总平均值'])
    
    # 合并结果
    all_avg = pd.concat([total_avg, class_avg])
    
    # 添加汇总列
    # 全卷（所有分数的总和）
    all_avg.insert(0, '全卷', all_avg[score_columns].sum(axis=1).round(2))
    # 语文（语文学科分数的总和），这里直接用相同的分数列求和，因为语文卷子只包含语文题
    # 如果未来需要区分不同学科，可传入 chinese_cols 单独计算，此处保持与原有逻辑一致
    all_avg.insert(1, '语文', all_avg[score_columns].sum(axis=1).round(2))
    
    return all_avg


# 生成学科分析（原 generate_chinese_analysis）
def generate_subject_analysis(data, subject_cols, subject_name='语文'):
    # 按班级分组计算
    class_analysis = data.groupby('行政班级')[subject_cols].mean().round(2)
    class_analysis = class_analysis.sort_index()
    
    # 全年级
    total_analysis = data[subject_cols].mean().round(2)
    total_analysis = pd.DataFrame([total_analysis], index=['总平均值'])
    
    # 合并
    all_analysis = pd.concat([total_analysis, class_analysis])
    
    # 添加总分列
    all_analysis.insert(0, f'{subject_name}成绩', all_analysis[subject_cols].sum(axis=1).round(2))
    
    return all_analysis


# 色阶效果
def add_color_scale(ws, start_col, end_col, start_row, end_row):
    try:
        range_str = f'{start_col}{start_row}:{end_col}{end_row}'
        color_scale_rule = ColorScaleRule(
            start_type='min', start_color='FFF16B61',
            mid_type='percentile', mid_value=50, mid_color='FFF7E98F',
            end_type='max', end_color='FF64BC7B'
        )
        ws.conditional_formatting.add(range_str, color_scale_rule)
    except Exception as e:
        print(f"添加色阶效果时出错: {str(e)}")


# 保存结果到Excel文件
def save_results_to_excel(all_avg, subject_analysis, output_file, data, subject_name='语文'):
    wb = _create_workbook(all_avg, subject_analysis, data, subject_name)
    wb.save(output_file)


# 保存结果到字节流（用于Web应用）
def save_results_to_excel_bytes(all_avg, subject_analysis, data, subject_name='语文'):
    wb = _create_workbook(all_avg, subject_analysis, data, subject_name)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# 创建工作簿的内部函数
def _create_workbook(all_avg, subject_analysis, data, subject_name='语文'):
    wb = Workbook()
    # 删除默认创建的空白 sheet
    wb.remove(wb.active)
    
    # 定义通用边框
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # ---------- 工作表1：学科分析 ----------
    ws_analysis = wb.create_sheet(title=f"{subject_name}学科分析")
    
    # 写入表头
    header_row = ['班级'] + list(subject_analysis.columns)
    ws_analysis.append(header_row)
    
    # 写入数据
    for idx, row in subject_analysis.iterrows():
        data_row = [idx] + list(row.values)
        ws_analysis.append(data_row)
    
    # 设置样式：表头
    for cell in ws_analysis[1]:
        cell.fill = PatternFill(start_color="FFCCCCCC", end_color="FFCCCCCC", fill_type="solid")
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # 数据区域居中和边框
    for row in ws_analysis.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
    
    # 为数据区域添加色阶效果（按列比较）
    if subject_analysis.shape[0] > 0 and subject_analysis.shape[1] > 0:
        start_row = 2
        end_row = start_row + len(subject_analysis) - 1
        for col_idx in range(subject_analysis.shape[1]):
            col_num = col_idx + 2  # B列开始（第2列）
            col_letter = ''
            temp = col_num
            while temp > 0:
                temp, remainder = divmod(temp - 1, 26)
                col_letter = chr(65 + remainder) + col_letter
            add_color_scale(ws_analysis, col_letter, col_letter, start_row, end_row)
    
    # 调整列宽和行高
    for column in ws_analysis.columns:
        ws_analysis.column_dimensions[column[0].column_letter].width = 8
    for row in ws_analysis.iter_rows():
        ws_analysis.row_dimensions[row[0].row].height = 14

    # ---------- 班级单独工作表 ----------
    classes = sorted(data['行政班级'].unique())
    
    # 提取分数列（与总分析保持一致）
    score_cols = get_score_columns(data)
    
    for cls in classes:
        class_data = data[data['行政班级'] == cls].copy()
        
        # 构建基础列
        base_columns = ['姓名', '班级']
        
        # 确保班级列存在
        if '行政班级' in class_data.columns:
            class_data['班级'] = class_data['行政班级']
        if '姓名' not in class_data.columns:
            class_data['姓名'] = ''
        
        # 确保全卷和语文列存在并计算
        # 如果原始数据没有“全卷”列（比如已被重命名），则根据分数列求和生成
        if '全卷' not in class_data.columns:
            class_data['全卷'] = class_data[score_cols].sum(axis=1).round(2)
        if '语文' not in class_data.columns:
            class_data['语文'] = class_data[score_cols].sum(axis=1).round(2)  # 对语文来说与全卷相同
        
        base_columns.extend(['全卷', '语文'])
        
        # 写入表头
        header_row = base_columns + score_cols
        ws_class = wb.create_sheet(title=f"{cls}")
        ws_class.append(header_row)
        
        # 写入数据行
        for _, student in class_data.iterrows():
            row_data = []
            for col in base_columns:
                row_data.append(student[col] if col in student else '')
            for col in score_cols:
                row_data.append(student[col] if col in student else '')
            ws_class.append(row_data)
        
        # 设置表头样式
        for cell in ws_class[1]:
            cell.fill = PatternFill(start_color="FFCCCCCC", end_color="FFCCCCCC", fill_type="solid")
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # 设置数据区域样式及零分标记
        base_col_count = len(base_columns)
        for row_idx in range(2, ws_class.max_row + 1):
            for col_idx in range(1, ws_class.max_column + 1):
                cell = ws_class.cell(row=row_idx, column=col_idx)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border
                # 为小题列中的0分单元格标红
                if col_idx > base_col_count and cell.value == 0:
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        # 设置列宽和行高
        for column in ws_class.columns:
            ws_class.column_dimensions[column[0].column_letter].width = 8
        for row in ws_class.iter_rows():
            ws_class.row_dimensions[row[0].row].height = 14

    return wb


# 主函数
def main():
    print("正在读取数据...")
    # 假设数据文件名为 data.xls，可自行修改
    data = read_excel_file('data.xls')
    print(f"成功读取 {len(data)} 条记录")
    
    # 提取小题列（自动去除答案列和固定列）
    score_cols = get_score_columns(data)
    print(f"识别到语文小题 {len(score_cols)} 个: {score_cols}")
    
    # 计算班级平均分
    print("正在计算班级平均分...")
    class_averages = calculate_class_averages(data, score_cols)
    print("班级平均分计算完成")
    
    # 生成学科分析
    print("正在生成语文学科分析...")
    chinese_analysis = generate_subject_analysis(data, score_cols, subject_name='语文')
    
    # 保存结果
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = f'分析结果_{timestamp}.xlsx'
    save_results_to_excel(class_averages, chinese_analysis, output_file, data, subject_name='语文')
    print(f"分析结果已保存到 '{output_file}'")


if __name__ == "__main__":
    main()